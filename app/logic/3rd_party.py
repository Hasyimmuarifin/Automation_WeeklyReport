from openpyxl import load_workbook

def move_data(file_a_path, file_b_path, sheet_a, sheet_b):
    """
    Move Plan and Actual data from file A to file B with flexible settings.
    If a value is missing, it will be replaced with 0.

    Args:
        file_a_path (str): Path to the source Excel file (file A).
        file_b_path (str): Path to the destination Excel file (file B).
        sheet_a (str): Sheet name in file A to read data from.
        sheet_b (str): Sheet name in file B to write data to.
    """

    # Open workbook A and select the sheet to read data from
    wb_a = load_workbook(file_a_path)
    ws_a = wb_a[sheet_a]

    # Open workbook B and select the sheet to write data to
    wb_b = load_workbook(file_b_path)
    ws_b = wb_b[sheet_b]

    # === General configuration mapping source rows to target start rows ===
    plans_and_actuals = [
        {'source_row_plan': 5, 'target_start_row': 3},
        {'source_row_plan': 6, 'target_start_row': 15},
        {'source_row_plan': 7, 'target_start_row': 27},
        {'source_row_plan': 8, 'target_start_row': 39},
        {'source_row_plan': 9, 'target_start_row': 51},
        {'source_row_plan': 10, 'target_start_row': 63},
        {'source_row_plan': 11, 'target_start_row': 75},
        {'source_row_plan': 12, 'target_start_row': 87},
        {'source_row_plan': 13, 'target_start_row': 111},
        {'source_row_plan': 14, 'target_start_row': 99},
        {'source_row_plan': 15, 'target_start_row': 123},
        {'source_row_plan': 16, 'target_start_row': 135},
        {'source_row_plan': 17, 'target_start_row': 147},
        {'source_row_plan': 18, 'target_start_row': 159},
        {'source_row_plan': 19, 'target_start_row': 171},
        {'source_row_plan': 20, 'target_start_row': 183},
        {'source_row_plan': 21, 'target_start_row': 219},
        {'source_row_plan': 22, 'target_start_row': 195},
        {'source_row_plan': 24, 'target_start_row': 243},
        {'source_row_plan': 25, 'target_start_row': 207},
    ]

    # Columns in file A that contain Plan and Actual data, alternating columns
    columns_in_a_plan = ['D', 'F', 'H', 'J', 'L', 'N', 'P', 'R', 'T', 'V', 'X', 'Z']
    columns_in_a_actual = ['E', 'G', 'I', 'K', 'M', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AA']

    # Target columns in file B for Plan and Actual respectively
    target_column_plan = 'E'
    target_column_actual = 'F'

    # === Processing PLAN data: copying from source rows and columns to target cells ===
    for item in plans_and_actuals:
        target_row = item['target_start_row']  # Start row for target sheet
        for col_a in columns_in_a_plan:
            # Read the data from the source cell
            column_data = ws_a[f"{col_a}{item['source_row_plan']}"].value
            # Write the data or 0 if data is None to the target cell
            ws_b[f"{target_column_plan}{target_row}"] = column_data if column_data is not None else 0
            target_row += 1  # Move to the next row in target

    # === Processing ACTUAL data similarly ===
    for item in plans_and_actuals:
        target_row = item['target_start_row']
        for col_a in columns_in_a_actual:
            column_data = ws_a[f"{col_a}{item['source_row_plan']}"].value
            ws_b[f"{target_column_actual}{target_row}"] = column_data if column_data is not None else 0
            target_row += 1

    # Save the updated workbook B to persist changes
    wb_b.save(file_b_path)