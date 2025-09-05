import openpyxl
import pandas as pd
import json
import os

# Class to help with file path management
class ResourceHelper:
    @staticmethod
    def get_path(relative_path: str) -> str:
        base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, relative_path)

# Get the path to the JSON configuration file
file_path_json = ResourceHelper.get_path('../config/inputan.json')

# Read the JSON configuration file
with open(file_path_json, 'r') as file:
    json_data = json.load(file)

# ======== Path to Excel files ========
data_final_path = json_data["final_file"]
data_summary_path = json_data["summary_file"]

# ======== Load data from file B ========
data_summary = pd.read_excel(data_summary_path, sheet_name='ITM Summary', header=json_data["header_month1"])
data_summary.columns = data_summary.columns.str.strip()

# Debug
print("Column names in file B:", data_summary.columns.tolist())

# ======== Load file A and select the sheet ========
wb = openpyxl.load_workbook(data_final_path)
ws = wb['ITM Summary']

# Mapping Column
columns_to_update = {
    'No.': 'A', 
    'Month': 'C', 
    'Company': 'D', 
    'Name of Vessel': 'F',
    'Buyer': 'G', 
    'End user': 'H', 
    'Load Port': 'I',
    'ETA/ATA': 'J',
    'ETB': 'L', 
    'ETD': 'N', 
    'Total': 'BJ', 
    'Lay': 'BK', 
    'can': 'BM',
    '%': 'BO', 
    'Status': 'BP', 
    'TM (AR)': 'BR', 
    'M (AD)': 'BS',
    'ASH (AD)': 'BT', 
    'ASH (AR)': 'BU', 
    'TS (AD)': 'BV', 
    'TS (AR)': 'BW',
    'CV (AD)': 'BX', 
    'CV (AR)': 'BY', 
    'CV (NAR)': 'BZ'
}

start_row = 4
end_row = start_row + json_data["data_count_month1"] - 1

# ======== Fill standard columns ========
for col_name, excel_col in columns_to_update.items():
    if col_name not in data_summary.columns:
        print(f"⚠️ Column '{col_name}' not found in file B.")
        continue

    for i, value in enumerate(data_summary[col_name].iloc[: end_row - start_row + 1]):
        ws[f"{excel_col}{start_row + i}"] = value

# ======== Format date function ========
def convert_to_date_format(date_value):
    """Convert date ke format 'd.Mmm' (contoh: 5.Sep)."""
    try:
        if pd.isna(date_value):
            return None

        if isinstance(date_value, str):
            # Format dd/mm
            if '/' in date_value:
                day, month = date_value.split('/')
                date_converted = pd.to_datetime(
                    f"2024-{month}-{day}", format="%Y-%m-%d", errors="coerce"
                )
            else:
                date_converted = pd.to_datetime(date_value, errors="coerce")
        else:
            date_converted = pd.to_datetime(date_value, errors="coerce")

        if pd.isna(date_converted):
            return None
        return f"{date_converted.day}.{date_converted.strftime('%b')}"
    except Exception as e:
        print(f"⚠️ Error converting '{date_value}': {e}")
        return None

# ======== Format ETA/ATA, ETB, ETD ========
for row in range(start_row, end_row + 1):
    for src_col, tgt_col in zip(['J', 'L', 'N'], ['K', 'M', 'O']):
        val = ws[f"{src_col}{row}"].value
        if val:
            ws[f"{tgt_col}{row}"] = convert_to_date_format(val)

# ======== Format Lay and Can ========
for row in range(start_row, end_row + 1):
    for src_col, tgt_col in zip(['BK', 'BM'], ['BL', 'BN']):
        val = ws[f"{src_col}{row}"].value
        if val:
            ws[f"{tgt_col}{row}"] = convert_to_date_format(val)

# ======== No Mahakam based on Load Port ========
no_mahakam = 1
for row in range(start_row, end_row + 1):
    load_port = ws[f"I{row}"].value
    if load_port == "BoCT":
        ws[f"B{row}"] = 0
    else:
        ws[f"B{row}"] = no_mahakam
        no_mahakam += 1

# ======== Type of Shipment based on Name of Vessel ========
for row in range(start_row, end_row + 1):
    name_vessel = ws[f"F{row}"].value
    if isinstance(name_vessel, str):
        vessel_upper = name_vessel.upper()
        if vessel_upper.startswith("MV"):
            ws[f"E{row}"] = "Vessel"
        elif vessel_upper.startswith("BG"):
            ws[f"E{row}"] = "Direct Shipment"
        elif "DUMP TRUCK" in vessel_upper:
            ws[f"E{row}"] = "Dump Truck"

# ======== Fill columns for BoCT only ========

for i, row_data in data_summary.iloc[: end_row - start_row + 1].iterrows():
    row = start_row + i
    load_port = ws[f"I{row}"].value
    if load_port != "BoCT":
        continue

    mapping_boCT = {
        'P': 'IMM-WB.HCV.LS',
        'Q': 'IMM-WB.MCV.HS',
        'R': 'IMM-EB.MCV.LS',
        'S': 'IMM-EB.MCV.MS',
        'T': 'IMM-EB.MCV.HS',
        'U': 'TCM.HCV.LS',
        'V': 'TCM.HCV.HS',
        'W': 'TCM.LCV.MS.HA',
        'X': 'BEK.MCV.LS',
        'Y': 'BEK.HCV.MS',
        'Z': 'JBG',
        'AA': 'GPK',
        'AB': 'TIS',

        #Third Party
        'AC': 'EBH.HCV',
        'AE': 'KMIA.MCV.LS',
        'AF': 'BBE.MCV',
        'AG': 'MBL.MCV',
        'AH': 'MBL.56.MCV',
        'AJ': 'EMJ.MCV',
        'AL': 'KBM.MCV',
        'AM': 'IKJ.MCV',
        'AN': 'MKE.LCV',
        'AQ': 'KJA.LCV.LS',
        'AS': 'DMP.LCV',
        'AT': 'MCM.LCV',
        'AU': 'BMM.LCV',
        'AW': 'BBA.LCV',
        'AX': 'MML.LCV',
        'BA': 'KPM.LCV',
        'BB': 'KJM.LCV',
        'BF': 'BUM.LCV',
        'BG': 'BISM.LCV',
    }

    for excel_col, col_name in mapping_boCT.items():
        if col_name in data_summary.columns:
            ws[f"{excel_col}{row}"] = row_data[col_name]

# ======== Save workbook ========
wb.save(data_final_path)
wb.close()
print("Excel file has been updated and saved.")

print("The columns have been successfully updated and saved back to the same file... :)")

# Main function placeholder
def main():
    print("ongoing_month processing")

# Run main if script executed directly
if __name__ == "__main__":
    main()