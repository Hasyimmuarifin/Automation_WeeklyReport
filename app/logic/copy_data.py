from openpyxl import load_workbook
import json
import os

class ResourceHelper:
    @staticmethod
    def get_path(relative_path: str) -> str:
        """
        Get the absolute path of a file relative to the current script's directory.

        Args:
            relative_path (str): The relative path to the file.

        Returns:
            str: The absolute path to the file.
        """
        # Get the directory of the current script and join it with the relative path
        base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, relative_path)

# Get the path to the JSON configuration file
file_path_json = ResourceHelper.get_path('../config/inputan.json')

# Read the JSON configuration file
with open(file_path_json, 'r') as file:
    json_data = json.load(file)

# Extract source and output file paths from the JSON data
source_file = json_data["summary_file"]
output_file = json_data["final_file"]
sheet_name = 'ITM Summary'  # Specify the sheet name to work with

# Get the header row and maximum number of data rows to copy
header_row = json_data["header_month1"]
max_row = json_data["data_count_month1"]

# Get the selected week for data extraction
selected_week = json_data["selected_week"]

# Mapping of columns for Penalty data based on the selected week
week_column_map_penalty = {
    'W0': 'AKC', 'W1': 'AKD', 'W2': 'AKE', 'W3': 'AKF', 'W4': 'AKG', 'W5': 'AKH'
}

# Mapping of columns for Demurrage data based on the selected week
week_column_map_demurrage = {
    'W0': 'AKK', 'W1': 'AKL', 'W2': 'AKM', 'W3': 'AKN', 'W4': 'AKO', 'W5': 'AKP'
}

# Function to copy weekly data to the output worksheet
def copy_column_data(ws_src, ws_out, week_key, col_map, col_output_index, label):
    """
    Copy data from the source worksheet to the output worksheet based on the selected week.

    Args:
        ws_src: Source worksheet object.
        ws_out: Output worksheet object.
        week_key (str): The key representing the selected week.
        col_map (dict): Mapping of week keys to column letters.
        col_output_index (int): The column index in the output worksheet to copy data to.
        label (str): A label for logging purposes.
    """
    # Validate if the week key exists in the column mapping
    if week_key not in col_map:
        raise ValueError(f"{label}: Week '{week_key}' not valid. Choose from {list(col_map.keys())}")
    
    # Get the column letter corresponding to the week key
    col_letter = col_map[week_key]
    
    # Validate the header to ensure it matches the expected week key
    header_check = ws_src[f"{col_letter}{header_row + 1}"].value
    if header_check != week_key:
        raise ValueError(f"{label}: Validation failed. Cell contents {col_letter}{header_row + 1} = '{header_check}', should be '{week_key}'")
    
    # Copy data from the source worksheet to the output worksheet
    start_row = header_row + 2  # Start copying data from the row after the header
    for i in range(max_row):
        value = ws_src[f"{col_letter}{start_row + i}"].value  # Get the value from the source cell
        if value is not None:  # Only copy non-empty values
            ws_out.cell(row=4 + i, column=col_output_index).value = value  # Write to the output cell
    
    print(f"{label} Week '{week_key}' (column {col_letter}) successfully copied to the index column {col_output_index}.")

# Function to copy total values and convert them to negative
def copy_total_value(ws_src, ws_out, week_key, col_map, output_col_index, label, offset):
    """
    Copy total values from the source worksheet to the output worksheet and convert them to negative.

    Args:
        ws_src: Source worksheet object.
        ws_out: Output worksheet object.
        week_key (str): The key representing the selected week.
        col_map (dict): Mapping of week keys to column letters.
        output_col_index (int): The column index in the output worksheet to copy the total value to.
        label (str): A label for logging purposes.
        offset (int): The row offset to locate the total value.
    """
    # Validate if the week key exists in the column mapping
    if week_key not in col_map:
        raise ValueError(f"{label}: Week '{week_key}' not valid.")
    
    # Get the column letter for the week and calculate the row number for total value
    col_letter = col_map[week_key]
    row_index = header_row + offset + max_row

    # Read the value from the source worksheet
    value = ws_src[f"{col_letter}{row_index}"].value

    # Check if the value is a string with parentheses, indicating a negative number
    if isinstance(value, str) and value.startswith("(") and value.endswith(")"):
        cleaned_value = value.strip("()")
        try:
            # Attempt to convert the cleaned value to a float
            # Note: This will raise an error if the value contains a comma
            formatted_value = -round(abs(float(cleaned_value)), 2)
            ws_out.cell(row=4, column=output_col_index).value = formatted_value
            print(f"{label} from {col_letter}{row_index} = '{value}' copied as '{formatted_value}' to the index column {output_col_index}.")
        except ValueError:
            # Value is not a valid number
            print(f"{label} from {col_letter}{row_index} not copied because the value in parentheses is not a valid number: '{value}'")

    # If the value is an integer or float, copy it as-is
    elif isinstance(value, (int, float)):
        formatted_value = round(value, 2)
        ws_out.cell(row=4, column=output_col_index).value = formatted_value
        print(f"{label} from {col_letter}{row_index} = '{value}' copied as '{formatted_value}' to the index column {output_col_index}.")

    # If the value is a string without parentheses, try to convert it to a float and copy it
    elif isinstance(value, str):
        try:
            # Note: This will also raise an error if the string uses a comma instead of a dot
            formatted_value = round(float(value), 2)
            ws_out.cell(row=4, column=output_col_index).value = formatted_value
            print(f"{label} from {col_letter}{row_index} = '{value}' copied as '{formatted_value}' to the index column {output_col_index}.")
        except ValueError:
            print(f"{label} from {col_letter}{row_index} not copied because the value is not a valid number: '{value}'")

    # If the value is neither numeric nor a numeric string, it will not be copied
    else:
        print(f"{label} from {col_letter}{row_index} not copied because the value is not a number: '{value}'")


# Main Function to perform copying operations
def main():
    print("Start the Excel file customization process...")
    
    # Load the source and output Excel workbooks
    wb_source = load_workbook(source_file, data_only=True)
    ws_source = wb_source[sheet_name]
    wb_output = load_workbook(output_file)
    ws_output = wb_output[sheet_name]

    # Copy weekly Penalty data to column 81 (CC)
    copy_column_data(ws_source, ws_output, selected_week, week_column_map_penalty, 81, "Penalty")

    # Copy weekly Demurrage data to column 89 (CK)
    copy_column_data(ws_source, ws_output, selected_week, week_column_map_demurrage, 89, "Demurrage")

    # Copy total Penalty BOCT to column 94 (CP)
    copy_total_value(ws_source, ws_output, selected_week, week_column_map_penalty, 94, "Penalty BOCT", offset=3)

    # Copy total Penalty Mahakam to column 95 (CQ)
    copy_total_value(ws_source, ws_output, selected_week, week_column_map_penalty, 95, "Penalty Mahakam", offset=4)

    # Copy total Demurrage BOCT to column 96 (CR)
    copy_total_value(ws_source, ws_output, selected_week, week_column_map_demurrage, 96, "Demurrage BOCT", offset=3)

    # Copy total Demurrage Mahakam to column 97 (CS)
    copy_total_value(ws_source, ws_output, selected_week, week_column_map_demurrage, 97, "Demurrage Mahakam", offset=4)

    # Save the output workbook with the applied changes
    wb_output.save(output_file)
    print("Output file is saved successfully.")

if __name__ == "__main__":
    main()