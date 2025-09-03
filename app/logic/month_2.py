import openpyxl
import pandas as pd
import json
import os

# Class to help with file path management
class ResourceHelper:
    @staticmethod
    def get_path(relative_path: str) -> str:
        """
        Get the absolute path of a file relative to the current script's directory.

        Args:
            relative_path (str): The relative path to the file.

        Returns:
            str: The absolute path to the file.
        """
        # Get the directory of the current script and join it with the relative path
        base_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_path, relative_path)

# Get the path to the JSON configuration file
file_path_json = ResourceHelper.get_path('../config/inputan.json')

# Read the JSON configuration file
with open(file_path_json, 'r') as file:
    json_data = json.load(file)

# Extract paths for the Excel files from the JSON data
data_final_path = json_data["final_file"]
data_summary_file = json_data["summary_file"]

# Read data from the summary Excel file, specifically from the 'ITM Summary' sheet
# Start reading from the specified header row
data_summary = pd.read_excel(data_summary_file, sheet_name='ITM Summary', header=json_data["header_month2"])

# Clean up column names by stripping excess whitespace
data_summary.columns = data_summary.columns.str.strip()

# Display the column names from the summary file for debugging purposes
print("Column names in file B:", data_summary.columns.tolist())

# Load the workbook from the final data file
wb = openpyxl.load_workbook(data_final_path)
ws = wb['Month 2']  # Change to the appropriate sheet name as needed

# Column Mapping
columns_to_update = {
    "No.": "A",
    "Month": "C",
    "Company": "D",
    "Name of Vessel": "F",
    "Buyer": "G",
    "End user": "H",
    "Load Port": "I",
    "ETA/ATA": "J",
    "ETB": "K",
    "ETD": "L",
    "Total": "BD",
    "Lay": "BE",
    "can": "BF",
    "%": "BG",
    "Status": "BH",

    "TM (AR)": "BJ",
    "M (AD)": "BK",
    "ASH (AD)": "BL",
    "ASH (AR)": "BM",
    "TS (AD)": "BN",
    "TS (AR)": "BO",
    "CV (AD)": "BP",
    "CV (AR)": "BQ",
    "CV (NAR)": "BR"
}

# Fill in data in the final Excel file from the summary data
start_row = 4  # The first row of data in the final file
num_rows_b = len(data_summary)  # Number of rows in the summary data
data_count = start_row + json_data["data_count_month2"]  # Calculate the total number of rows to fill

# Loop through each column to update and fill in the data
for column_name, excel_column in columns_to_update.items():
    for index in range(num_rows_b):
        if index < (data_count - start_row):  # Limit the number of rows being filled
            try:
                value_to_write = data_summary[column_name].iloc[index]  # Get the value from the summary data
                ws[f'{excel_column}{start_row + index}'] = value_to_write  # Write the value to the final file
                print(f"Copy From {column_name} To {excel_column}{start_row + index}: {value_to_write}")  # Debugging output
            except KeyError:
                print(f"Column '{column_name}' Not Found in file B.")  # Handle missing columns
            except IndexError:
                print(f"Insufficient data in column '{column_name}' for index {index}.")  # Handle index errors

# Function to convert date values to a specific format (DD.MMM)
def convert_to_date_format(date_value):
    """
    Convert a date value to the format DD.MMM (e.g., 13.Apr).

    Args:
        date_value: The date value to convert.

    Returns:
        str: The formatted date string or None if conversion fails.
    """
    try:
        if isinstance(date_value, str) and '/' in date_value:
            day, month = date_value.split('/')  # Split the string into day and month
            date_converted = pd.to_datetime(f'2024-{month}-{day}', format='%Y-%m-%d')  # Convert to datetime
        else:
            date_converted = pd.to_datetime(date_value)

        day = date_converted.day
        month_abbr = date_converted.strftime('%b')

        return f"{day}.{month_abbr}"
    except Exception as e:
        print(f"Error while converting value {date_value}: {e}")
        return None

# Initialize Mahakam number starting at 1
no_mahakam = 1

# Loop to fill "No Mahakam" column based on Load Port values
for row in range(start_row, min(start_row + num_rows_b, data_count)):
    load_port = ws[f'I{row}'].value  # Column I is Load Port
    if load_port == 'BoCT':
        ws[f'B{row}'] = 0  # Set No Mahakam to 0 for BoCT
    elif load_port in ['SMD Anc', 'Bunyut']:
        ws[f'B{row}'] = no_mahakam  # Assign current Mahakam number
        no_mahakam += 1
    else:
        ws[f'B{row}'] = no_mahakam  # Assign and increment by default
        no_mahakam += 1

# Loop to determine Type of Shipment based on Name of Vessel
for row in range(start_row, num_rows_b + 1):
    name_of_vessel = ws[f'F{row}'].value  # Column F is Name of Vessel
    print(f"Row {row}, Name of Vessel: {name_of_vessel}")  # Debug info

    # Convert name to uppercase for uniform evaluation
    if isinstance(name_of_vessel, str):
        name_of_vessel_upper = name_of_vessel.upper()
        if name_of_vessel_upper.startswith('MV'):
            ws[f'E{row}'] = 'Vessel'
        elif name_of_vessel_upper.startswith('BG'):
            ws[f'E{row}'] = 'Direct Shipment'
        elif 'DUMP TRUCK' in name_of_vessel_upper:
            ws[f'E{row}'] = 'Dump Truck'
        else:
            ws[f'E{row}'] = None
    else:
        ws[f'E{row}'] = None

# Save the workbook back to the file
wb.save(data_final_path)
wb.close()

print("The columns have been successfully updated and saved back to the same file... :)")

# Main function placeholder
def main():
    print("month 2 processing")

# Run main function if script executed directly
if __name__ == "__main__":
    main()