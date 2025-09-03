from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re
import copy
import json
from pathlib import Path
import os

class ResourceHelper:
    @staticmethod
    def get_path(relative_path: str) -> Path:
        """
        Resolve an absolute path for a file relative to the current script.

        Args:
            relative_path (str): Relative file path from the script location.

        Returns:
            Path: Absolute Path object to the target file.
        """
        # Use pathlib to calculate base path and append relative path
        base_path = Path(__file__).parent
        return base_path / relative_path

# ────────────────── Reading Configuration from JSON ────────────────────────
CFG_PATH = ResourceHelper.get_path('../config/inputan.json')

# Load configuration data from JSON file with UTF-8 encoding
cfg = json.loads(CFG_PATH.read_text(encoding="utf-8"))

# Extract target Excel file path from the configuration
file_path = cfg["final_file"]

# Mapping sheet names to tuples of (number of data rows expected, table name in Excel)
data_counts_and_tables: dict[str, tuple[int, str]] = {
    "ITM Summary": (cfg["data_count_month1"], "TableOngoing"),
    "Month 1":     (cfg["data_count_month1"], "TableMonth1"),
    "Month 2":     (cfg["data_count_month2"], "TableMonth2"),
    "Month 3":     (cfg["data_count_month3"], "TableMonth3"),
    "Month 4":     (cfg["data_count_month4"], "TableMonth4"),
}

# Load the Excel workbook specified in the configuration
wb = load_workbook(file_path)

# Iterate over all sheets listed in the mapping dictionary
for sheet_name, (data_count, table_name) in data_counts_and_tables.items():

    # Skip this sheet if data count is zero (means no update needed)
    if data_count == 0:
        print(f"Skip '{sheet_name}', data does not change.")
        continue

    # Access the worksheet object by sheet name
    ws = wb[sheet_name]

    # Locate the desired table object by matching the table name
    table = next((tbl for tbl in ws.tables.values() if tbl.name == table_name), None)
    if table is None:
        # If the table is not found, report and skip this sheet
        print(f"Table '{table_name}' not found in sheet '{sheet_name}'.")
        continue

    # Parse the table reference range (e.g. 'A3:F20') into start and end coordinates
    start_cell, end_cell = table.ref.split(":")

    # Extract starting row number (adding 1 to skip header row in table)
    start_row = int(re.match(r"[A-Za-z]+(\d+)", start_cell).group(1)) + 1

    # Extract ending column as an index number (1-based)
    end_col = column_index_from_string(re.match(r"([A-Za-z]+)", end_cell).group())

    # Extract ending row number of the existing table range
    end_row = int(re.match(r"[A-Za-z]+(\d+)", end_cell).group(1))

    # Calculate the current number of data rows inside the table (excluding header)
    current_rows = end_row - start_row + 1

    # Initialize variable to track new last row of the table after modification
    new_end_row = end_row

    # ── Add Rows if needed to match desired data count ────────────────────────
    if current_rows < data_count:
        rows_to_add = data_count - current_rows  # Number of rows to add

        # For each row to add
        for i in range(rows_to_add):
            # For each column in the table
            for col_idx in range(1, end_col + 1):
                # Reference the last existing row cell to copy style from
                source_cell = ws.cell(end_row, col_idx)
                # Target cell for the new row
                target_cell = ws.cell(end_row + 1 + i, col_idx)
                target_cell.value = None  # Initialize new cell value to None

                # Copy the cell style properties if source cell has any style
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.number_format = source_cell.number_format
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.alignment = copy.copy(source_cell.alignment)

        # Update the new last row number of the table
        new_end_row = end_row + rows_to_add
        print(f"{rows_to_add} line(s) added in '{sheet_name}' (until row {new_end_row}).")

    # ── Remove Rows if excess to match desired data count ────────────────────
    elif current_rows > data_count:
        rows_to_remove = current_rows - data_count  # Number of rows to remove

        # Delete excess rows from the bottom of the table's data section
        ws.delete_rows(end_row - rows_to_remove + 1, rows_to_remove)

        # Update the new last row number of the table
        new_end_row = end_row - rows_to_remove
        print(f"{rows_to_remove} line(s) removed from '{sheet_name}'.")

    # ── If current rows already matches desired data count ───────────────────
    else:
        print(f"'{sheet_name}' is up to date with {data_count} line(s).")

    # Update the table reference to reflect the changed data range
    table.ref = f"{start_cell}:{ws.cell(row=new_end_row, column=end_col).coordinate}"

# Save all changes back to the Excel file
wb.save(file_path)

# Close the workbook explicitly to free any resources
wb.close()

print("The file was successfully customized and resaved.")

def main():
    # Main entry point; intentionally left empty when run as a module
    pass

# If this script is executed directly, call main()
if __name__ == "__main__":
    main()